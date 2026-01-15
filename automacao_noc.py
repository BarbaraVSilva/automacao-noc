import pyperclip
import openpyxl
import time
import os
import keyboard
import ctypes
import configparser
from copy import copy
from datetime import datetime

# --- CARREGAMENTO DE CONFIGURAÇÕES COM INTERAÇÃO INICIAL ---
def carregar_e_configurar():
    config = configparser.ConfigParser()
    arquivo_conf = 'config.ini'
    
    # Padrão inicial
    padrao = {
        'GERAL': {
            'nome_analista': 'Analista NOC',
            'arquivo_excel': 'NOC - BASE BI 2024.xlsx'
        },
        'OPCOES': {
            'criticidade': 'High, Critical, Disaster',
            'origem': 'Zabbix, New Relic, Dynatrace, ZENA, Grafana'
        }
    }

    # Se não existe o arquivo, cria o padrão
    if not os.path.exists(arquivo_conf):
        config.read_dict(padrao)
        with open(arquivo_conf, 'w', encoding='utf-8') as f:
            config.write(f)
    else:
        config.read(arquivo_conf, encoding='utf-8')
    
    # Verifica se o nome ainda é o padrão para pedir o nome do analista
    nome_atual = config.get('GERAL', 'nome_analista', fallback='Analista NOC')
    if nome_atual == 'Analista NOC':
        # Traz o terminal para frente para o usuário ver a pergunta
        try:
            hwnd = ctypes.WinDLL('kernel32').GetConsoleWindow()
            if hwnd: ctypes.WinDLL('user32').SetForegroundWindow(hwnd)
        except: pass

        print("\n" + "═"*55)
        print(" ✨ CONFIGURAÇÃO INICIAL DO ASSISTENTE")
        print(" " + "═"*55)
        novo_nome = input(" 👉 Digite seu nome completo (Analista): ").strip()
        
        if novo_nome:
            nome_atual = novo_nome
            config.set('GERAL', 'nome_analista', nome_atual)
            with open(arquivo_conf, 'w', encoding='utf-8') as f:
                config.write(f)
            print(f"\n✅ Perfeito, {nome_atual}! Seu nome foi salvo.")
            time.sleep(1.5)
            os.system('cls' if os.name == 'nt' else 'clear')

    return config, nome_atual

# Inicialização das variáveis globais
config_data, NOME_ANALISTA = carregar_e_configurar()

NOME_ARQUIVO_EXCEL = config_data.get('GERAL', 'arquivo_excel', fallback='NOC - BASE BI 2024.xlsx')
criticidade_str = config_data.get('OPCOES', 'criticidade', fallback='High, Critical, Disaster')
OPCOES_CRITICIDADE = [i.strip() for i in criticidade_str.split(',')]
origem_str = config_data.get('OPCOES', 'origem', fallback='Zabbix, New Relic, Dynatrace, ZENA, Grafana')
OPCOES_ORIGEM = [i.strip() for i in origem_str.split(',')]

# Caminhos Dinâmicos
DIRETORIO_ATUAL = os.path.dirname(os.path.abspath(__file__))
CAMINHO_EXCEL = os.path.join(DIRETORIO_ATUAL, NOME_ARQUIVO_EXCEL)
ARQUIVO_LOG = os.path.join(DIRETORIO_ATUAL, "historico_automacao_noc.txt")

CAMPOS_PARA_COPIAR = [
    "INCIDENTE (Col A)", "APLICAÇÃO AFETADA (Col B)", 
    "HORÁRIO DO ALERTA (Col D)", "EQUIPE ACIONADA (Col E)", 
    "ANALISTA ACIONADO (Col F)"
]

def focar_terminal():
    try:
        hwnd = ctypes.WinDLL('kernel32').GetConsoleWindow()
        if hwnd:
            ctypes.WinDLL('user32').ShowWindow(hwnd, 9)
            ctypes.WinDLL('user32').SetForegroundWindow(hwnd)
    except: pass

def registrar_log(mensagem):
    timestamp = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    try:
        with open(ARQUIVO_LOG, "a", encoding="utf-8") as f:
            f.write(f"[{timestamp}] {mensagem}\n")
    except: pass

def buscar_primeira_linha_vazia(ws):
    linha = 2 
    while ws.cell(row=linha, column=1).value is not None:
        linha += 1
    return linha

def salvar_no_excel(dados_dicio):
    try:
        if not os.path.exists(CAMINHO_EXCEL):
            print(f"\n❌ ERRO: Arquivo '{NOME_ARQUIVO_EXCEL}' não encontrado na pasta.")
            return

        try:
            f = open(CAMINHO_EXCEL, "a")
            f.close()
        except IOError:
            print("\n⚠️ ERRO: A planilha está aberta! Feche-a e tente novamente.")
            return

        wb = openpyxl.load_workbook(CAMINHO_EXCEL)
        ws = wb.active
        proxima_linha = buscar_primeira_linha_vazia(ws)
        linha_ref = proxima_linha - 1 if proxima_linha > 2 else 2
        
        mapeamento = {
            "A": dados_dicio["incidente"], "B": dados_dicio["app"],
            "C": dados_dicio["data_auto"], "D": dados_dicio["horario_copiado"],
            "E": dados_dicio["equipe"],    "F": dados_dicio["analista_ac"],
            "G": dados_dicio["criticidade"],"H": dados_dicio["origem"],
            "J": "Não",                    "L": NOME_ANALISTA
        }
        
        for col, valor in mapeamento.items():
            nova_cel = ws[f"{col}{proxima_linha}"]
            nova_cel.value = valor
            cel_ref = ws[f"{col}{linha_ref}"]
            if cel_ref.has_style:
                nova_cel.font = copy(cel_ref.font)
                nova_cel.border = copy(cel_ref.border)
                nova_cel.fill = copy(cel_ref.fill)
                nova_cel.number_format = copy(cel_ref.number_format)
                nova_cel.alignment = copy(cel_ref.alignment)
        
        wb.save(CAMINHO_EXCEL)
        wb.close()
        print(f"\n✅ DADOS SALVOS NA LINHA {proxima_linha}!")
        registrar_log(f"Sucesso: {dados_dicio['incidente']} salvo por {NOME_ANALISTA}")
    except Exception as e:
        print(f"\n❌ ERRO CRÍTICO AO SALVAR: {e}")

def menu_selecao(titulo, opcoes):
    focar_terminal()
    print(f"\n--- {titulo} ---")
    for i, opt in enumerate(opcoes, 1):
        print(f"{i}. {opt}")
    while True:
        esc = input(f"Selecione (1-{len(opcoes)}): ")
        if esc.isdigit() and 1 <= int(esc) <= len(opcoes):
            return opcoes[int(esc)-1]

def modo_captura():
    data_hoje = datetime.now().strftime("%d/%m/%Y")
    print("\n" + "═"*55)
    print(f" 🚀 INICIANDO CAPTURA | {data_hoje}")
    print(" [ESC] Cancelar captura")
    print("═"*55)
    
    respostas = {"data_auto": data_hoje}
    chaves = ["incidente", "app", "horario_copiado", "equipe", "analista_ac"]
    
    for i, campo in enumerate(CAMPOS_PARA_COPIAR):
        confirmado = False
        while not confirmado:
            pyperclip.copy("")
            print(f"👉 Copie: {campo}", end="\r")
            
            item = ""
            while item == "":
                if keyboard.is_pressed('esc'): 
                    print("\n🛑 Captura cancelada."); return
                item = pyperclip.paste().strip()
                time.sleep(0.1)
            
            print(f"✅ Lido: {item[:35]}... | [F10] OK | [F8] REPETIR")
            while True:
                if keyboard.is_pressed('f10'):
                    respostas[chaves[i]] = item
                    confirmado = True
                    time.sleep(0.3); break
                if keyboard.is_pressed('f8'):
                    time.sleep(0.3); break
                if keyboard.is_pressed('esc'): 
                    print("\n🛑 Captura cancelada."); return

    respostas["criticidade"] = menu_selecao("CRITICIDADE", OPCOES_CRITICIDADE)
    respostas["origem"] = menu_selecao("ORIGEM", OPCOES_ORIGEM)
    salvar_no_excel(respostas)
    print("\n🏁 Aguardando novo chamado (F9)...")

if __name__ == "__main__":
    os.system("title AUTOMACAO NOC")
    print(f"=====================================================")
    print(f" ASSISTENTE NOC ATIVO - ANALISTA: {NOME_ANALISTA}")
    print(f"=====================================================")
    print(" [F9]  INICIAR NOVO CHAMADO")
    print(" [End] FECHAR PROGRAMA")
    print(f"-----------------------------------------------------")
    
    keyboard.add_hotkey('f9', modo_captura)
    keyboard.wait('end')